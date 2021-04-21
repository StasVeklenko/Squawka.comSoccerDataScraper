import openpyxl, os
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from collections import OrderedDict

class spreadsheet:
    def __init__(self, workbook, worksheet, keep_vba=False): #delete name of workbook as there is only 1 being initialized; only leave name of workbook, and all the worksheets in a dictionary
        self.worksheets = {}
        if keep_vba:
            self.workbook = openpyxl.load_workbook(workbook, keep_vba=True)
        else:
            self.workbook = openpyxl.load_workbook(workbook)
        self.name = workbook    
        self.worksheets[worksheet] = self.workbook.get_sheet_by_name(worksheet)
    
    def get_data(self, worksheet, row_start, column_start, incr_step, number_of_incrs, incr_along):
        data = []
        for i in range(number_of_incrs): #might have to change it to set_data way of incrementing 
            if (incr_along == "row"):
                data.append(self.worksheets[worksheet].cell(row=row_start,column=column_start+i*incr_step).value)
            elif (incr_along == "column"):
                data.append(self.worksheets[worksheet].cell(row=row_start+i*incr_step,column=column_start).value)        
        return data
            
    def set_data(self, data, worksheet, row_start, column_start, incr_step = 1, incr_along = "row", workbook = "", keep_vba=False):   
        sheet = self.worksheets[worksheet]
        if type(data) == list:
            for i in range(len(data)):
                if (incr_along == "row"):
                    sheet.cell(row=row_start,column=column_start+incr_step*i).value = str(data[i])
                elif (incr_along == "column"):
                    sheet.cell(row=row_start+incr_step*i,column=column_start).value = str(data[i]) 
        elif type(data) == OrderedDict:
            i = 0
            for key in data.keys():
                if (incr_along == "row"):
                    sheet.cell(row=row_start,column=column_start+incr_step*i).value = str(data[key]) 
                elif (incr_along == "column"):
                    sheet.cell(row=row_start+incr_step*i,column=column_start).value = str(data[key])
                i += 1
        elif type(data) == int or type(data) == str:
            sheet.cell(row=row_start,column=column_start).value = data           
        '''
        if not workbook:
            workbook_name = self.name
        else:
            workbook_name = workbook
            
        if not keep_vba:
            self.workbook.save(workbook_name)
        else:
            self.workbook.save(workbook_name, keep_vba=True)
        '''
    def save(self, workbook="", keep_vba=False):
        if not workbook:
            workbook_name = self.name
        else:
            workbook_name = workbook
    
        if not keep_vba:
            self.workbook.save(workbook_name)
        else:
            self.workbook.save(workbook_name, keep_vba=True)        